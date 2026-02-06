import os

from reader import ExcelReader
from codegen import CodeGen
from outputs import ROutputs
from variable import VariableConverter
import openpyxl
import copy
import unittest
import rpy2.robjects as robjects
import hashlib
import re
import numpy as np

def file_hash(path):
    hasher = hashlib.sha256()
    with open(path, 'rb') as file:
        for chunk in iter(lambda: file.read(4096), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


class TestWorkbook1(unittest.TestCase):
    """

    """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excel workbook/test_workbook_1.xlsx"
        self.rpath = "test/excel workbook/test_workbook_1_output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = ["PSA"]
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)
        outputs = ROutputs(varconverter, "", [], [], [], [], [], BCEA=False)
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode,outputs, "", self.rpath)
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.culledcode = copy.deepcopy(codegen.orderedcode)
        codegen.none_strip()
        codegen.generate_code(writeoutputs=False)

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

    def test_deterministic_conversion(self):
        """
        Given no changes to the Excel file, returned files should always be the same.
        We can test this using a cryptographic hash. If the contents of the files are different then the conversion
        has not generated the same file.
        """

        hashOriginal = file_hash("test/excel workbook/test_workbook_1.R")
        hashGenerated = file_hash(self.rpath)
        self.assertEqual(hashOriginal,hashGenerated, "Hash comparison failed - converter update has changed deterministic output\n")

    def test_programatic_regression(self):
        """
        While the converted file may be correct, the individual functions called may change definition when called in R
        or via incompatible changes from Excel. This test catches when these regressions may occur
        """

        self.assertAlmostEqual(self.globalenv[f'Sheet1_A1'][0], 3,
                               msg=f"Sheet1_A1 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Sheet1_A2'][0], 3,
                               msg=f"Sheet1_A2 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Sheet1_A3'][0], 5,
                               msg=f"Sheet1_A3 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Sheet1_A4'][0], 8,
                               msg=f"Sheet1_A4 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Sheet1_D2'][0], 5,
                               msg=f"Sheet1_D2 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Sheet1_D3'][0], 6,
                               msg=f"Sheet1_D3 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Sheet1_D4'][0], 7,
                               msg=f"Sheet1_D4 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Sheet1_D5'][0], 8,
                               msg=f"Sheet1_D5 ... [FAIL]\n")
        self.assertEqual(self.globalenv[f'Sheet1_C1'][0], "sdf",
                               msg=f"Sheet1_D5 ... [FAIL]\n")

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()
        os.remove("test/excel workbook/test_workbook_1_output.R")


class TestWorkbook2(unittest.TestCase):
    """

    """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excel workbook/test_workbook_2.xlsx"
        self.rpath = "test/excel workbook/test_workbook_2_output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = ["PSA"]
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)
        outputs = ROutputs(varconverter, "", [], [], [], [], [], BCEA=False )
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode,outputs , "", self.rpath)
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.culledcode = copy.deepcopy(codegen.orderedcode)
        codegen.none_strip()
        codegen.generate_code(writeoutputs=False)

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

    def test_deterministic_conversion(self):
        """
        Given no changes to the Excel file, returned files should always be the same.
        We can test this using a cryptographic hash. If the contents of the files are different then the conversion
        has not generated the same file.
        """

        hashOriginal = file_hash("test/excel workbook/test_workbook_2.R")
        hashGenerated = file_hash(self.rpath)
        self.assertEqual(hashOriginal,hashGenerated, "Hash comparison failed - converter update has changed deterministic output\n")

    def test_programatic_regression(self):
        """
        While the converted file may be correct, the individual functions called may change definition when called in R
        or via incompatible changes from Excel. This test catches when these regressions may occur
        """

        self.assertAlmostEqual(self.globalenv[f'willingness_to_pay'][0], 20000,
                               msg=f"willingness_to_pay (Frontend_E5) ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Frontend_E7'][0], 4600,
                               msg=f"Frontend_E7 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Frontend_E8'][0], 1540,
                               msg=f"Frontend_E8 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_E5'][0], 100,
                               msg=f"Engine_E5 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_E6'][0], 19.2,
                               msg=f"Engine_E6 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_E7'][0], 383900,
                               msg=f"Engine_E7 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_F5'][0], 560,
                               msg=f"Engine_F5 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_F6'][0], 19.3,
                               msg=f"Engine_F6 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_F7'][0], 385440,
                               msg=f"Engine_F7 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_G5'][0], 460,
                               msg=f"Engine_G5 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_G6'][0], 0.1,
                               msg=f"Engine_G6 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_G7'][0], 1540,
                               msg=f"Engine_G7 ... [FAIL]\n")

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()
        os.remove("test/excel workbook/test_workbook_2_output.R")

class TestWorkbook3(unittest.TestCase):
    """

    """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excel workbook/test_workbook_3.xlsx"
        self.rpath = "test/excel workbook/test_workbook_3_output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = ["PSA"]
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)
        outputs = ROutputs(varconverter, "", [], [], [], [], [], BCEA=False)
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode,outputs , "", self.rpath)
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.culledcode = copy.deepcopy(codegen.orderedcode)
        codegen.none_strip()
        codegen.generate_code(writeoutputs=False)

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

    def test_deterministic_conversion(self):
        """
        Given no changes to the Excel file, returned files should always be the same.
        We can test this using a cryptographic hash. If the contents of the files are different then the conversion
        has not generated the same file.
        """

        hashOriginal = file_hash("test/excel workbook/test_workbook_3.R")
        hashGenerated = file_hash(self.rpath)
        self.assertEqual(hashOriginal,hashGenerated, "Hash comparison failed - converter update has changed deterministic output\n")

    def test_programatic_regression(self):
        """
        While the converted file may be correct, the individual functions called may change definition when called in R
        or via incompatible changes from Excel. This test catches when these regressions may occur
        """

        self.assertAlmostEqual(self.globalenv[f'willingness_to_pay'][0], 20000,
                               msg=f"willingness_to_pay (Frontend_E5) ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Frontend_E8'][0], 2731.83216783217,
                               msg=f"Frontend_E8 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Frontend_E9'][0], 12346.74,
                               msg=f"Frontend_E9 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_E5'][0], 705,
                               msg=f"Engine_E5 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_E6'][0], 18.82,
                               msg=f"Engine_E6 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_E7'][0], 375695,
                               msg=f"Engine_E7 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_F5'][0], 2658.26,
                               msg=f"Engine_F5 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_F6'][0], 19.535,
                               msg=f"Engine_F6 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_F7'][0], 388041.74,
                               msg=f"Engine_F7 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_G5'][0], 1953.26,
                               msg=f"Engine_G5 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_G6'][0], 0.715,
                               msg=f"Engine_G6 ... [FAIL]\n")
        self.assertAlmostEqual(self.globalenv[f'Engine_G7'][0], 12346.74,
                               msg=f"Engine_G7 ... [FAIL]\n")

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()
        os.remove("test/excel workbook/test_workbook_3_output.R")

class TestWorkbook4(unittest.TestCase):
    """

    """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excel workbook/test_workbook_4.xlsm"
        self.rpath = "test/excel workbook/test_workbook_4_output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = ["PSA"]
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)

        costs = [('Engine', 'E5:F5')]
        effs = [('Engine', 'E6:F6')]
        otherPSA = [('Engine', 'E7:F7')]
        outputs = ROutputs(varconverter, "test/excel workbook/", otherPSA, costs, effs,
                           ["Asprin", "Warfarin"], [('Frontend','E5')], BCEA=False)
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode, outputs, "test/excel workbook/", "test_workbook_4_output.R")
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.cyclic_prune()
        codegen.generate_code(writeoutputs=True)

        with open(self.rpath,"r+") as f:
            file_data = f.read()
            file_data = re.sub("numberOfRuns = 1000", "numberOfRuns = 100000", file_data)
            file_data = re.sub("converter_validate = TRUE", "converter_validate = FALSE", file_data)
            f.seek(0)
            f.write(file_data)
            f.truncate()

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

        self.psaData = np.loadtxt("test/excel workbook/psa.txt", skiprows=1)
        os.remove("test/excel workbook/psa.txt")

    def test_deterministic_conversion(self):
        """
        Given no changes to the Excel file, returned files should always be the same.
        We can test this using a cryptographic hash. If the contents of the files are different then the conversion
        has not generated the same file.
        """

        hashOriginal = file_hash("test/excel workbook/test_workbook_4.R")
        hashGenerated = file_hash(self.rpath)
        self.assertEqual(hashOriginal,hashGenerated, "Hash comparison failed - converter update has changed deterministic output\n")

    def test_programatic_regression(self):
        """
        While the converted file may be correct, the individual functions called may change definition when called in R
        or via incompatible changes from Excel. This test catches when these regressions may occur
        """

        EngineE5, EngineF5, EngineE6, EngineF6, EngineE7, EngineF7 = np.average(self.psaData, axis=0)

        avg_inc_costs = EngineF5 - EngineE5
        avg_inc_effs = EngineF6 - EngineE6
        avg_inc_nb = EngineF7 - EngineE7
        print(avg_inc_costs)
        print(avg_inc_effs)
        print(avg_inc_nb)

        # Assume statistical accuracy to 1%
        self.assertAlmostEqual(-1425.659922, avg_inc_costs/avg_inc_effs,
                               msg=f"Frontend_E9 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.",delta = abs(-1425.659922)*0.01)
        self.assertAlmostEqual(-30965.96106, avg_inc_nb,
                               msg=f"Frontend_E9 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta = abs(-30965.96106)*0.01)

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()
        os.remove("test/excel workbook/test_workbook_4_output.R")


class TestTwoStateMarkovV2(unittest.TestCase):
    """

    """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excel workbook/Two states Markov model_v0.2.xlsm"
        self.rpath = "test/excel workbook/Two states Markov model_v0.2_output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = ["PSA"]
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)

        costs = [('Results', 'F12:F13')]
        effs = [('Results', 'G12:G13')]
        otherPSA = [('Results', 'H12:H13')]
        outputs = ROutputs(varconverter, "test/excel workbook/", otherPSA, costs, effs,
                           ["Drug A", "Drug B"], [('Model settings','N13')], BCEA=False)
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode, outputs, "test/excel workbook/", "Two states Markov model_v0.2_output.R")
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.cyclic_prune()
        codegen.generate_code(writeoutputs=True)

        with open(self.rpath,"r+") as f:
            file_data = f.read()
            file_data = re.sub("numberOfRuns = 1000", "numberOfRuns = 100000", file_data)
            file_data = re.sub("converter_validate = TRUE", "converter_validate = FALSE", file_data)
            f.seek(0)
            f.write(file_data)
            f.truncate()

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

        self.psaData = np.loadtxt("test/excel workbook/psa.txt", skiprows=1)


    def test_deterministic_conversion(self):
        """
        Given no changes to the Excel file, returned files should always be the same.
        We can test this using a cryptographic hash. If the contents of the files are different then the conversion
        has not generated the same file.
        """

        hashOriginal = file_hash("test/excel workbook/Two states Markov model_v0.2.R")
        hashGenerated = file_hash(self.rpath)
        self.assertEqual(hashOriginal,hashGenerated, "Hash comparison failed - converter update has changed deterministic output\n")

    def test_programatic_regression(self):
        """
        While the converted file may be correct, the individual functions called may change definition when called in R
        or via incompatible changes from Excel. This test catches when these regressions may occur
        """

        avgResultsF12, avgResultsF13, avgResultsG12, avgResultsG13,avgResultsH12, avgResultsH13 = np.average(self.psaData, axis=0)

        # Assume statistical accuracy to 1%
        self.assertAlmostEqual(8132.14, avgResultsF12,
                               msg=f"Frontend_E9 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=8132.14 * 0.01)

        self.assertAlmostEqual(6867.66, avgResultsF13,
                               msg=f"Frontend_E9 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=6867.66 * 0.01)

        self.assertAlmostEqual(7.59, avgResultsG12,
                               msg=f"Frontend_E9 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=7.59 * 0.01)

        self.assertAlmostEqual(6.96, avgResultsG13,
                               msg=f"Frontend_E9 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=6.96* 0.01)

        self.assertAlmostEqual(143915.57, avgResultsH12,
                               msg=f"Frontend_E9 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta = 143915.57*0.01)

        self.assertAlmostEqual(131413.43, avgResultsH13,
                               msg=f"Frontend_E9 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=131413.43*0.01)

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()
        os.remove("test/excel workbook/psa.txt")
        os.remove("test/excel workbook/Two states Markov model_v0.2_output.R")

class TestHIPSdemo(unittest.TestCase):
    """

    """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excel workbook/HIPS Markov model demo.xlsm"
        self.rpath = "test/excel workbook/HIPS Markov model demo_output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = ["DSA results", "PSA results"]
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)

        costs = [('Summary results', 'D6:D9')]
        effs = [('Summary results', 'E6:E9')]
        otherPSA = []
        outputs = ROutputs(varconverter, "test/excel workbook/", otherPSA, costs, effs,
                           ["Cemented", "Uncemented", "Hybrid", "Reverse Hybrid"], [('Setup and run','D14')], BCEA=False)
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode, outputs, "test/excel workbook/", "HIPS Markov model demo_output.R")
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.cyclic_prune()
        codegen.generate_code(writeoutputs=True)

        with open(self.rpath,"r+") as f:
            file_data = f.read()
            file_data = re.sub("numberOfRuns = 1000", "numberOfRuns = 3000", file_data)
            file_data = re.sub("converter_validate = TRUE", "converter_validate = FALSE", file_data)
            f.seek(0)
            f.write(file_data)
            f.truncate()

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

        self.psaData = np.loadtxt("test/excel workbook/psa.txt", skiprows=1)

    def test_deterministic_conversion(self):
        """
        Given no changes to the Excel file, returned files should always be the same.
        We can test this using a cryptographic hash. If the contents of the files are different then the conversion
        has not generated the same file.
        """

        hashOriginal = file_hash("test/excel workbook/HIPS Markov model demo.R")
        hashGenerated = file_hash(self.rpath)
        self.assertEqual(hashOriginal,hashGenerated, "Hash comparison failed - converter update has changed deterministic output\n")

    def test_programatic_regression(self):
        """
        While the converted file may be correct, the individual functions called may change definition when called in R
        or via incompatible changes from Excel. This test catches when these regressions may occur
        """

        avgVals= np.average(self.psaData, axis=0)

        # Assume statistical accuracy to 1%
        self.assertAlmostEqual(1973.04, avgVals[0],
                               msg=f"PSA results C22 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=1973.04 * 0.01)

        self.assertAlmostEqual(3181.16, avgVals[1],
                               msg=f"PSA results G22 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=3181.16 * 0.01)

        self.assertAlmostEqual(2562.80, avgVals[2],
                               msg=f"PSA results K22 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=2562.80 * 0.01)

        self.assertAlmostEqual(2495.46, avgVals[3],
                               msg=f"PSA results O22 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=2495.46 * 0.01)

        self.assertAlmostEqual(15.04, avgVals[4],
                               msg=f"PSA results E22 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=15.04 * 0.01)

        self.assertAlmostEqual(15.11, avgVals[5],
                               msg=f"PSA results I22 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=15.11 * 0.01)

        self.assertAlmostEqual(15.08, avgVals[6],
                               msg=f"PSA results M22 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=15.08 * 0.01)

        self.assertAlmostEqual(15.07, avgVals[7],
                               msg=f"PSA results O22 ... [FAIL]\n"
                                   f"- Potentially failed due to random sampling.\n"
                                   f"Run at least 3 times if it failed.\n "
                                   f"If fails 2 out of 3 times check program.", delta=15.07 * 0.01)


    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()
        os.remove("test/excel workbook/psa.txt")
        os.remove("test/excel workbook/HIPS Markov model demo_output.R")


if __name__ == '__main__':
    unittest.main()
