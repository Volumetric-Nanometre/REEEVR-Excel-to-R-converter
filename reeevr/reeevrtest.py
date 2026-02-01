from reader import ExcelReader
from codegen import CodeGen
from outputs import ROutputs
from variable import VariableConverter
import openpyxl
import copy
import unittest
import rpy2.robjects as robjects

@unittest.skip
class SimpleFunctionTests(unittest.TestCase):
    """
    Unit tests covering the functions treated as simple transforms in formula.py
    These functions should all work exactly the same in R as in Excel (including order of arguments).
    """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excelunit/simple-tests.xlsx"
        self.rpath = "test/excelunit/simple-test-output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = []
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)
        outputs = ROutputs(varconverter, "", "", [], [], [], [], [])
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode,outputs , codefile=self.rpath)
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.culledcode = copy.deepcopy(codegen.orderedcode)
        codegen.none_strip()
        codegen.generate_code(writeoutputs=False)

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv
    #
    # Single value function tests
    #
    def test_full_excel(self):
        for i in range(3, 10):
            function = self.globalenv[f'Sheet1_B{i}'][0]
            self.assertAlmostEqual(self.globalenv[f'Sheet1_I{i}'][0], self.globalenv[f'Sheet1_O{i}'][0],
                                   msg=f"{function} - Sheet1_I{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_J{i}'][0], self.globalenv[f'Sheet1_P{i}'][0],
                                   msg=f"{function} - Sheet1_J{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_K{i}'][0], self.globalenv[f'Sheet1_Q{i}'][0],
                                   msg=f"{function} - Sheet1_K{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_L{i}'][0], self.globalenv[f'Sheet1_R{i}'][0],
                                   msg=f"{function} - Sheet1_L{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_M{i}'][0], self.globalenv[f'Sheet1_S{i}'][0],
                                   msg=f"{function} - Sheet1_M{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_N{i}'][0], self.globalenv[f'Sheet1_T{i}'][0],
                                   msg=f"{function} - Sheet1_N{i}\n ... [FAIL]")
            print(f"{function} ... [PASS]")

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()

class AdaptiveFunctionTests(unittest.TestCase):
    """
       Unit tests covering the functions treated as simple transforms in formula.py
       These functions require an adjustment to the function name.
       e.g. LN and LOG work differently in Excel vs R.

       Test suite explicity loads an Excel file containing all functions and test cases.
       File is then converted using the standard conversion process (minus culling).
       File is then ran in R using rpy2, this generates an active R environment.
       We then interrogate the R environment and compare to the expected value stored in Excel.
       """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excelunit/adaptive-tests.xlsx"
        self.rpath = "test/excelunit/adaptive-test-output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = []
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)
        outputs = ROutputs(varconverter, "", "", [], [], [], [], [])
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode,outputs , codefile=self.rpath)
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.culledcode = copy.deepcopy(codegen.orderedcode)
        codegen.none_strip()
        codegen.generate_code(writeoutputs=False)

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

    def test_adaptive_excel(self):
        for i in range(3, 17):
            function = self.globalenv[f'Sheet1_B{i}'][0]
            self.assertAlmostEqual(self.globalenv[f'Sheet1_I{i}'][0], self.globalenv[f'Sheet1_O{i}'][0],
                                   msg=f"{function} - Sheet1_I{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_J{i}'][0], self.globalenv[f'Sheet1_P{i}'][0],
                                   msg=f"{function} - Sheet1_J{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_K{i}'][0], self.globalenv[f'Sheet1_Q{i}'][0],
                                   msg=f"{function} - Sheet1_K{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_L{i}'][0], self.globalenv[f'Sheet1_R{i}'][0],
                                   msg=f"{function} - Sheet1_L{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_M{i}'][0], self.globalenv[f'Sheet1_S{i}'][0],
                                   msg=f"{function} - Sheet1_M{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_N{i}'][0], self.globalenv[f'Sheet1_T{i}'][0],
                                   msg=f"{function} - Sheet1_N{i}\n ... [FAIL]")
            print(f"{function} ... [PASS]")

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()

class DistFunctionTests(unittest.TestCase):
    """
       Unit tests covering the functions treated as simple transforms in formula.py
       These functions require an adjustment to the function name.
       e.g. LN and LOG work differently in Excel vs R.

       Test suite explicity loads an Excel file containing all functions and test cases.
       File is then converted using the standard conversion process (minus culling).
       File is then ran in R using rpy2, this generates an active R environment.
       We then interrogate the R environment and compare to the expected value stored in Excel.
       """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excelunit/reeevr-dist-tests.xlsx"
        self.rpath = "test/excelunit/reeevr-dist-tests-output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = []
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)
        outputs = ROutputs(varconverter, "", "", [], [], [], [], [])
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode,outputs , codefile=self.rpath)
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.culledcode = copy.deepcopy(codegen.orderedcode)
        codegen.none_strip()
        codegen.generate_code(writeoutputs=False)

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

    def test_dist_excel(self):
        for i in range(3, 71):
            function = self.globalenv[f'Sheet1_B{i}'][0]
            self.assertAlmostEqual(self.globalenv[f'Sheet1_I{i}'][0], self.globalenv[f'Sheet1_O{i}'][0],
                                   msg=f"{function} - Sheet1_I{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_J{i}'][0], self.globalenv[f'Sheet1_P{i}'][0],
                                   msg=f"{function} - Sheet1_J{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_K{i}'][0], self.globalenv[f'Sheet1_Q{i}'][0],
                                   msg=f"{function} - Sheet1_K{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_L{i}'][0], self.globalenv[f'Sheet1_R{i}'][0],
                                   msg=f"{function} - Sheet1_L{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_M{i}'][0], self.globalenv[f'Sheet1_S{i}'][0],
                                   msg=f"{function} - Sheet1_M{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_N{i}'][0], self.globalenv[f'Sheet1_T{i}'][0],
                                   msg=f"{function} - Sheet1_N{i}\n ... [FAIL]")
            print(f"{function} ... [PASS]")

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()

class ArrayValidationTests(unittest.TestCase):
    """
       Unit tests covering the functions treated as simple transforms in formula.py
       These functions require an adjustment to the function name.
       e.g. LN and LOG work differently in Excel vs R.

       Test suite explicity loads an Excel file containing all functions and test cases.
       File is then converted using the standard conversion process (minus culling).
       File is then ran in R using rpy2, this generates an active R environment.
       We then interrogate the R environment and compare to the expected value stored in Excel.
       """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excelunit/array-validation-tests.xlsx"
        self.rpath = "test/excelunit/array-validation-tests-output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = []
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)
        outputs = ROutputs(varconverter, "", "", [], [], [], [], [])
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode,outputs , codefile=self.rpath)
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.culledcode = copy.deepcopy(codegen.orderedcode)
        codegen.none_strip()
        codegen.generate_code(writeoutputs=False)

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

    def test_array_validation_excel(self):
        for i in range(13, 25):
            function = self.globalenv[f'Sheet1_B{i}'][0]
            self.assertAlmostEqual(self.globalenv[f'Sheet1_C{i}'][0], self.globalenv[f'Sheet1_R{i}'][0],
                                   msg=f"{function} - Sheet1_C{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_D{i}'][0], self.globalenv[f'Sheet1_R{i}'][1],
                                   msg=f"{function} - Sheet1_D{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_E{i}'][0], self.globalenv[f'Sheet1_R{i}'][2],
                                   msg=f"{function} - Sheet1_E{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_F{i}'][0], self.globalenv[f'Sheet1_R{i}'][3],
                                   msg=f"{function} - Sheet1_F{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_G{i}'][0], self.globalenv[f'Sheet1_R{i}'][4],
                                   msg=f"{function} - Sheet1_G{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_H{i}'][0], self.globalenv[f'Sheet1_R{i}'][5],
                                   msg=f"{function} - Sheet1_H{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_I{i}'][0], self.globalenv[f'Sheet1_R{i}'][6],
                                   msg=f"{function} - Sheet1_I{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_J{i}'][0], self.globalenv[f'Sheet1_R{i}'][7],
                                   msg=f"{function} - Sheet1_J{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_K{i}'][0], self.globalenv[f'Sheet1_R{i}'][8],
                                   msg=f"{function} - Sheet1_J{i}\n ... [FAIL]")
            print(f"{function} ... [PASS]")

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()

class REEEVRFunctionTests(unittest.TestCase):
    """
       Unit tests covering the functions treated as simple transforms in formula.py
       These functions require an adjustment to the function name.
       e.g. LN and LOG work differently in Excel vs R.

       Test suite explicity loads an Excel file containing all functions and test cases.
       File is then converted using the standard conversion process (minus culling).
       File is then ran in R using rpy2, this generates an active R environment.
       We then interrogate the R environment and compare to the expected value stored in Excel.
       """
    @classmethod
    def setUpClass(self):
        """
        Test suite explicity loads an Excel file containing all functions and test cases.
        File is then converted using the standard conversion process (minus culling).
        File is then ran in R using rpy2, this generates an active R environment.
        We then interrogate the R environment and compare to the expected value stored in Excel.
        """
        self.excelpath = "test/excelunit/reeevr-tests.xlsx"
        self.rpath = "test/excelunit/reeevr-test-output.R"
        workbook = openpyxl.load_workbook(self.excelpath)
        outputLang = 'R'
        ignoredsheets = []
        varconverter = VariableConverter(workbook, ignoredsheets, outputLang)
        outputs = ROutputs(varconverter, "", "", [], [], [], [], [])
        reader = ExcelReader(varconverter, workbook, outputLang, ignoredsheets)
        reader.read()
        codegen = CodeGen(varconverter, reader.unorderedcode,outputs , codefile=self.rpath)
        codegen.second_pass()
        codegen.order_code_snippets()
        codegen.culledcode = copy.deepcopy(codegen.orderedcode)
        codegen.none_strip()
        codegen.generate_code(writeoutputs=False)

        self.r_source = robjects.r['source']
        self.r_source(self.rpath)
        self.globalenv = robjects.globalenv

    def test_reeevr_excel(self):
        for i in range(3, 35):
            function = self.globalenv[f'Sheet1_B{i}'][0]
            self.assertAlmostEqual(self.globalenv[f'Sheet1_I{i}'][0], self.globalenv[f'Sheet1_O{i}'][0],
                                   msg=f"{function} - Sheet1_I{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_J{i}'][0], self.globalenv[f'Sheet1_P{i}'][0],
                                   msg=f"{function} - Sheet1_J{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_K{i}'][0], self.globalenv[f'Sheet1_Q{i}'][0],
                                   msg=f"{function} - Sheet1_K{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_L{i}'][0], self.globalenv[f'Sheet1_R{i}'][0],
                                   msg=f"{function} - Sheet1_L{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_M{i}'][0], self.globalenv[f'Sheet1_S{i}'][0],
                                   msg=f"{function} - Sheet1_M{i}\n ... [FAIL]")
            self.assertAlmostEqual(self.globalenv[f'Sheet1_N{i}'][0], self.globalenv[f'Sheet1_T{i}'][0],
                                   msg=f"{function} - Sheet1_N{i}\n ... [FAIL]")
            print(f"{function} ... [PASS]")

    @classmethod
    def tearDownClass(self):
        robjects.globalenv.clear()



if __name__ == '__main__':

    unittest.main()
